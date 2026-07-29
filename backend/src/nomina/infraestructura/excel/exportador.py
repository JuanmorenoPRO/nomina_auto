"""Exportador de liquidaciones a Excel con el formato de la planilla de la
contadora: una hoja por empleado con el desglose por concepto, más un resumen.
En la segunda quincena se agrega una hoja de APROPIACIONES."""

from __future__ import annotations

import re
from decimal import Decimal
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from nomina.aplicacion.casos_uso.liquidar_quincena import (
    LiquidacionEmpleado,
    LiquidacionQuincena,
)

_NEGRITA = Font(bold=True)
_TITULO = Font(bold=True, size=13)
_PESOS = "#,##0"
_BORDE_FINO = Border(bottom=Side(style="thin"))

_COLUMNAS_APROP = [
    ("EMPLEADO", None),
    ("CON AUXILIO", None),
    ("SIN AUXILIO", None),
    ("SENA", "aprop_sena"),
    ("ICBF", "aprop_icbf"),
    ("CAJA COMPENSACION", "aprop_caja_compensacion"),
    ("SALUD", "aprop_salud_total"),
    ("PENSION", "aprop_pension_total"),
    ("ARL", "aprop_arl"),
    ("VACACIONES", "aprop_vacaciones"),
    ("PRIMA", "aprop_prima"),
    ("CESANTIAS", "aprop_cesantias"),
    ("INTERESES", "aprop_intereses_cesantias"),
]

# Columnas que usan base CON AUXILIO (prima, cesantías); el resto usa SIN AUXILIO.
_BASE_CON_AUXILIO = {"aprop_prima", "aprop_cesantias"}


def _titulo_hoja(nombre: str) -> str:
    """Excel limita títulos a 31 caracteres y prohíbe []:*?/\\"""
    return re.sub(r"[\[\]:*?/\\]", "", nombre).strip()[:31] or "EMPLEADO"


def _encabezado(hoja: Worksheet, liq: LiquidacionQuincena) -> None:
    hoja["A1"] = liq.unidad.nombre
    hoja["A1"].font = _TITULO
    hoja["A2"] = f"NIT: {liq.unidad.nit}" if liq.unidad.nit else ""
    hoja["A3"] = (
        f"PERIODO DE PAGO: {liq.periodo.fecha_inicio:%d/%m/%Y} AL "
        f"{liq.periodo.fecha_fin:%d/%m/%Y}  (versión {liq.version})"
    )
    hoja["A3"].font = _NEGRITA


def _hoja_empleado(hoja: Worksheet, liq: LiquidacionQuincena, le: LiquidacionEmpleado) -> None:
    _encabezado(hoja, liq)
    hoja["A5"] = "EMPLEADO:"
    hoja["B5"] = le.empleado.nombre
    hoja["B5"].font = _NEGRITA
    hoja["D5"] = "CÉDULA:"
    hoja["E5"] = le.empleado.documento
    hoja["A6"] = "SALARIO BÁSICO MENSUAL:"
    hoja["C6"] = int(le.liquidacion.salario_mensual)
    hoja["C6"].number_format = _PESOS

    encabezados = ["CONCEPTO", "HORAS", "VALOR"]
    fila = 8
    for col, texto in enumerate(encabezados, start=1):
        celda = hoja.cell(row=fila, column=col, value=texto)
        celda.font = _NEGRITA
        celda.border = _BORDE_FINO

    for concepto in le.liquidacion.conceptos:
        fila += 1
        hoja.cell(row=fila, column=1, value=concepto.nombre)
        if concepto.minutos:
            celda_horas = hoja.cell(row=fila, column=2, value=float(round(concepto.horas, 2)))
            celda_horas.number_format = "0.00"
        hoja.cell(row=fila, column=3, value=int(concepto.valor)).number_format = _PESOS

    fila += 2
    hoja.cell(row=fila, column=1, value="TOTAL DEVENGADO").font = _NEGRITA
    celda_total = hoja.cell(row=fila, column=3, value=int(le.liquidacion.total_devengado))
    celda_total.font = _NEGRITA
    celda_total.number_format = _PESOS

    # Bloque de deducciones (si la unidad descuenta seguridad social o hay otras).
    if le.liquidacion.deducciones:
        fila += 2
        hoja.cell(row=fila, column=1, value="DEDUCCIONES").font = _NEGRITA
        for deduccion in le.liquidacion.deducciones:
            fila += 1
            hoja.cell(row=fila, column=1, value=deduccion.nombre)
            hoja.cell(row=fila, column=3, value=int(deduccion.valor)).number_format = _PESOS
        fila += 1
        hoja.cell(row=fila, column=1, value="TOTAL DEDUCCIONES").font = _NEGRITA
        celda_ded = hoja.cell(row=fila, column=3, value=int(le.liquidacion.total_deducciones))
        celda_ded.font = _NEGRITA
        celda_ded.number_format = _PESOS

    fila += 2
    hoja.cell(row=fila, column=1, value="VALOR A PAGAR").font = _TITULO
    celda_neto = hoja.cell(row=fila, column=3, value=int(le.liquidacion.neto_a_pagar))
    celda_neto.font = _TITULO
    celda_neto.number_format = _PESOS

    hoja.column_dimensions["A"].width = 42
    for col in "BCDE":
        hoja.column_dimensions[col].width = 14


def _hoja_resumen(hoja: Worksheet, liq: LiquidacionQuincena) -> None:
    _encabezado(hoja, liq)
    encabezados = ["EMPLEADO", "DOCUMENTO", "CARGO", "SALARIO MENSUAL",
                   "TOTAL DEVENGADO", "TOTAL DEDUCCIONES", "VALOR A PAGAR"]
    fila = 5
    for col, texto in enumerate(encabezados, start=1):
        celda = hoja.cell(row=fila, column=col, value=texto)
        celda.font = _NEGRITA
        celda.border = _BORDE_FINO

    for le in liq.por_empleado:
        fila += 1
        hoja.cell(row=fila, column=1, value=le.empleado.nombre)
        hoja.cell(row=fila, column=2, value=le.empleado.documento)
        hoja.cell(row=fila, column=3, value=le.empleado.cargo)
        hoja.cell(row=fila, column=4, value=int(le.liquidacion.salario_mensual)).number_format = _PESOS
        hoja.cell(row=fila, column=5, value=int(le.liquidacion.total_devengado)).number_format = _PESOS
        hoja.cell(row=fila, column=6, value=int(le.liquidacion.total_deducciones)).number_format = _PESOS
        hoja.cell(row=fila, column=7, value=int(le.liquidacion.neto_a_pagar)).number_format = _PESOS

    fila += 2
    hoja.cell(row=fila, column=1, value="TOTAL UNIDAD").font = _NEGRITA
    total_neto = sum((le.liquidacion.neto_a_pagar for le in liq.por_empleado), start=0)
    hoja.cell(row=fila, column=5, value=int(liq.total)).number_format = _PESOS
    celda = hoja.cell(row=fila, column=7, value=int(total_neto))
    celda.font = _NEGRITA
    celda.number_format = _PESOS
    hoja.cell(row=fila, column=5).font = _NEGRITA

    hoja.column_dimensions["A"].width = 34
    for col in range(2, 8):
        hoja.column_dimensions[get_column_letter(col)].width = 18
    hoja["A1"].alignment = Alignment(horizontal="left")


def _auxilio_de(le: LiquidacionEmpleado) -> Decimal:
    """Valor del auxilio de transporte en la liquidación (0 si no aplica)."""
    for c in le.liquidacion.conceptos:
        if c.codigo == "auxilio_transporte":
            return c.valor
    return Decimal(0)


def _redondear(v: Decimal) -> int:
    from decimal import ROUND_HALF_UP
    return int(v.quantize(Decimal("1"), rounding=ROUND_HALF_UP))


def _hoja_apropiaciones(
    hoja: Worksheet,
    segunda: LiquidacionQuincena,
    primera: LiquidacionQuincena,
    tasas: dict[str, Decimal],
) -> None:
    """Genera la hoja APROPIACIONES sumando las dos quincenas del mes."""
    hoja["A1"] = "APROPIACIONES SEGURIDAD SOCIAL"
    hoja["A1"].font = _TITULO
    hoja["A2"] = (
        f"{segunda.unidad.nombre}  —  "
        f"{primera.periodo.fecha_inicio:%d/%m/%Y} al {segunda.periodo.fecha_fin:%d/%m/%Y}"
    )
    hoja["A2"].font = _NEGRITA

    # Fila de encabezados (fila 4)
    fila_enc = 4
    for col, (nombre, _) in enumerate(_COLUMNAS_APROP, start=1):
        celda = hoja.cell(row=fila_enc, column=col, value=nombre)
        celda.font = _NEGRITA
        celda.border = _BORDE_FINO

    # Fila de factores (fila 5) — columnas D en adelante (índice 4+)
    fila_factor = 5
    for col, (_, codigo) in enumerate(_COLUMNAS_APROP, start=1):
        if codigo:
            celda = hoja.cell(row=fila_factor, column=col, value=float(tasas[codigo]))
            celda.number_format = "0.0000"

    # Construir mapa empleado_id → LiquidacionEmpleado para ambas quincenas
    mapa_primera: dict = {le.empleado.id: le for le in primera.por_empleado}
    mapa_segunda: dict = {le.empleado.id: le for le in segunda.por_empleado}
    empleados_ids = list({le.empleado.id for le in segunda.por_empleado} |
                         {le.empleado.id for le in primera.por_empleado})
    # Ordenar por nombre
    empleados_ids.sort(key=lambda eid: (
        mapa_segunda.get(eid) or mapa_primera.get(eid)
    ).empleado.nombre)

    fila = 5
    filas_datos: list[int] = []
    for emp_id in empleados_ids:
        le1 = mapa_primera.get(emp_id)
        le2 = mapa_segunda.get(emp_id)
        nombre_emp = (le2 or le1).empleado.nombre  # type: ignore[union-attr]

        dev1 = le1.liquidacion.total_devengado if le1 else Decimal(0)
        dev2 = le2.liquidacion.total_devengado if le2 else Decimal(0)
        aux1 = _auxilio_de(le1) if le1 else Decimal(0)
        aux2 = _auxilio_de(le2) if le2 else Decimal(0)

        con_auxilio = dev1 + dev2
        sin_auxilio = (dev1 - aux1) + (dev2 - aux2)

        fila += 1
        filas_datos.append(fila)
        hoja.cell(row=fila, column=1, value=nombre_emp)

        hoja.cell(row=fila, column=2, value=_redondear(con_auxilio)).number_format = _PESOS
        hoja.cell(row=fila, column=3, value=_redondear(sin_auxilio)).number_format = _PESOS

        cesantias_valor = Decimal(0)
        for col, (_, codigo) in enumerate(_COLUMNAS_APROP, start=1):
            if not codigo:
                continue
            tasa = tasas[codigo]
            if codigo == "aprop_intereses_cesantias":
                valor = cesantias_valor * tasa
            elif codigo in _BASE_CON_AUXILIO:
                valor = con_auxilio * tasa
            else:
                valor = sin_auxilio * tasa

            if codigo == "aprop_cesantias":
                cesantias_valor = valor

            hoja.cell(row=fila, column=col, value=_redondear(valor)).number_format = _PESOS

    # Fila de totales
    fila += 2
    hoja.cell(row=fila, column=1, value="TOTALES").font = _NEGRITA
    for col in range(2, len(_COLUMNAS_APROP) + 1):
        letra = get_column_letter(col)
        if filas_datos:
            primera_fila = filas_datos[0]
            ultima_fila = filas_datos[-1]
            celda = hoja.cell(
                row=fila, column=col,
                value=f"=SUM({letra}{primera_fila}:{letra}{ultima_fila})"
            )
        else:
            celda = hoja.cell(row=fila, column=col, value=0)
        celda.font = _NEGRITA
        celda.number_format = _PESOS

    hoja.column_dimensions["A"].width = 36
    for col in range(2, len(_COLUMNAS_APROP) + 1):
        hoja.column_dimensions[get_column_letter(col)].width = 16


def exportar_liquidacion_excel(
    liq: LiquidacionQuincena,
    primera_quincena: LiquidacionQuincena | None = None,
    tasas_apropiaciones: dict[str, Decimal] | None = None,
) -> bytes:
    libro = Workbook()

    # Hoja APROPIACIONES al frente si es segunda quincena con datos completos
    if primera_quincena is not None and tasas_apropiaciones is not None:
        aprop = libro.active
        aprop.title = "APROPIACIONES"
        _hoja_apropiaciones(aprop, liq, primera_quincena, tasas_apropiaciones)
        resumen = libro.create_sheet("RESUMEN")
    else:
        resumen = libro.active
        resumen.title = "RESUMEN"

    _hoja_resumen(resumen, liq)
    for le in liq.por_empleado:
        hoja = libro.create_sheet(_titulo_hoja(le.empleado.nombre))
        _hoja_empleado(hoja, liq, le)

    contenido = BytesIO()
    libro.save(contenido)
    return contenido.getvalue()
