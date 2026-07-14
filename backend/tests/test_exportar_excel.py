"""Exportación a Excel: el archivo que recibe la contadora."""

from io import BytesIO

from openpyxl import load_workbook


def _preparar_liquidacion(client) -> str:
    unidad = client.post(
        "/unidades", json={"nombre": "Edificio Thunapa P.H.", "nit": "800254433"}
    ).json()
    empleado = client.post(
        "/empleados",
        json={"unidad_id": unidad["id"], "nombre": "FREDY ALONSO HURTADO",
              "documento": "71712119", "cargo": "vigilante", "salario_base": 2_200_000},
    ).json()
    periodo = client.post(
        "/periodos", json={"fecha_inicio": "2026-06-16", "fecha_fin": "2026-06-30"}
    ).json()
    client.post(
        "/turnos",
        json={"empleado_id": empleado["id"], "fecha": "2026-06-28",
              "hora_inicio": "18:00", "hora_fin": "06:00"},
    )
    r = client.post(f"/periodos/{periodo['id']}/liquidar", json={"unidad_id": unidad["id"]})
    return r.json()["id"]


def test_descarga_excel_con_desglose(client):
    liquidacion_id = _preparar_liquidacion(client)

    r = client.get(f"/liquidaciones/{liquidacion_id}/excel")
    assert r.status_code == 200
    assert "spreadsheetml" in r.headers["content-type"]
    assert 'filename="liquidacion_Edificio_Thunapa_P.H._2026-06-16_2026-06-30_v1.xlsx"' in (
        r.headers["content-disposition"]
    )

    libro = load_workbook(BytesIO(r.content))
    assert libro.sheetnames == ["RESUMEN", "FREDY ALONSO HURTADO"]

    resumen = libro["RESUMEN"]
    assert resumen["A1"].value == "Edificio Thunapa P.H."
    assert resumen["A6"].value == "FREDY ALONSO HURTADO"
    assert resumen["E6"].value == 1_479_048  # total devengado del empleado

    hoja = libro["FREDY ALONSO HURTADO"]
    filas = {
        hoja.cell(row=f, column=1).value: hoja.cell(row=f, column=4).value
        for f in range(9, 15)
    }
    assert filas["TIEMPO ORDINARIO"] == 1_100_000
    assert filas["TIEMPO FESTIVO"] == 18_000
    assert filas["TIEMPO NOCTURNO DOMINICAL/FESTIVO"] == 236_500
    assert filas["AUXILIO DE TRANSPORTE"] == 124_548


def test_excel_de_liquidacion_inexistente(client):
    r = client.get("/liquidaciones/00000000-0000-0000-0000-000000000000/excel")
    assert r.status_code == 404
